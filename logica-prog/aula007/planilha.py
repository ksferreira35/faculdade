
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Criar um novo arquivo Excel
workbook = Workbook()
sheet = workbook.active
sheet.title = "Planilha Formatada"


# Definir estilos
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')


data_font = Font(name='Calibri', size=11)
data_alignment = Alignment(horizontal='left', vertical='center')


# Tamanho das colunas
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 30


# Escrever o cabeçalho com estilo
header_row = ["Nome", "Idade", "Profissão"]
for col_num, column_title in enumerate(header_row, 1):
   cell = sheet.cell(row=1, column=col_num, value=column_title)
   cell.font = header_font
   cell.fill = header_fill
   cell.alignment = header_alignment


# Escrever os dados com estilo
data = [
   ("Alice", 30, "Engenheira"),
   ("Bob", 25, "Designer"),
   ("Charlie", 35, "Analista de Dados")
]


for row_num, row_data in enumerate(data, 2):
   for col_num, cell_value in enumerate(row_data, 1):
       cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
       cell.font = data_font
       cell.alignment = data_alignment


# Salvar o arquivo Excel
workbook.save("planilha_formatada.xlsx")
print("Planilha 'planilha_formatada.xlsx' criada com formatação!")
