# pip install prettytable
from prettytable import PrettyTable
# pip install openpyxl
import openpyxl

tarefas = []

def incluir():
    print('--- Inclusão de Tarefa ---')
    tarefa = input('Digite a tarefa: ')
    prioridade = input('Digite a prioridade (1-Baixa/2-Média/3-Alta): ')

    txt_prio = ''

    if prioridade == '1':
        txt_prio = 'Baixa'

    elif prioridade == '2':
        txt_prio = 'Média'

    elif prioridade == '3':
        txt_prio = 'Alta'

    tupla = (tarefa, txt_prio)
    tarefas.append(tupla)
      


def listar():
    tabela = PrettyTable(["Número", "Tarefa", "Prioridade"])

    tabela.padding_width = 1
    numero = 0

    for tarefa, txt_prio in tarefas:
        tabela.add_row([numero, tarefa, txt_prio])

        numero += 1
        
    print(tabela)
    input("Digite [ENTER] para continuar")



def excluir():
    listar()
    num = int(input("Qual número da tarefa deseja apagar? "))
    print(f"\033[1;31mA opção {tarefas[num][0]} foi apagada!\033[m")
    tarefas.pop(num)


def gerar_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Planilha de Tarefas"

    # Tamanho das colunas
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15

    # Escrever o cabeçalho
    sheet.cell(row=1, column=1, value='Número')
    sheet.cell(row=1, column=2, value='Tarefa')
    sheet.cell(row=1, column=3, value='Prioridade')

    # Listar os dados
    linha = 2
    numero = 0
    for tarefa, txt_prio in tarefas:
        sheet.cell(row=linha, column=1, value=numero)
        sheet.cell(row=linha, column=2, value=tarefa)
        sheet.cell(row=linha, column=3, value=txt_prio)
        numero += 1
        linha += 1  # 👈 isso aqui tava faltando!

    # Salvar o arquivo Excel
    workbook.save("tarefas.xlsx")
    print("Planilha de tarefas gerada com sucesso!")
    input('Digite [ENTER] para continuar')



if __name__ == '__main__':
   # Menu
   terminar = False

   while terminar == False:
       print(f'''
{"---" * 7}
        MENU
1 - Incluir Tarefa
2 - Listar Tarefa
3 - Excluir Tarefas
4 - Gerar Excel
5 - Sair
{"---" * 7}''')

       opc = input('Qual a opção? ')
       if opc == '1':
           incluir()
       elif opc == '2':
           listar()
       elif opc == '3':
           excluir()
       elif opc == '4':
           gerar_excel()
       elif opc == '5':
           terminar = True

   print('Programa finalizado')